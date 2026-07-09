"""
Genba Management System — Import Seed Data Script.

Run this script to import Customers and Genbas from the Excel master data.
Usage:
    docker-compose exec api python -m app.scripts.import_seed_data /path/to/excel.xlsx
"""

import sys
import asyncio
import logging
import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.core.database import async_session_factory, engine
from app.modules.customer.models import CustomerModel
from app.modules.genba.models import GenbaModel
from app.modules.staff.models import StaffModel
from app.modules.worker.models import WorkerModel

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

async def import_data(file_path: str):
    logger.info(f"Loading Excel file: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "現場一覧" not in wb.sheetnames:
        logger.error("Sheet '現場一覧' not found.")
        return

    ws = wb["現場一覧"]
    
    try:
        async with async_session_factory() as session:
            # Keep track of customers created during this run to avoid duplicates
            customer_cache = {}

            # Cache all existing genba property names to avoid N+1 queries
            result_genba = await session.execute(select(GenbaModel.property_name))
            existing_genba_names = set(result_genba.scalars().all())

            # Rows start at 4 (1-indexed)
            count_customers = 0
            count_genbas = 0
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
                try:
                    customer_name = row[0]
                    property_name = row[10]
                    address = row[11]

                    if not property_name:
                        continue

                    customer_name_str = str(customer_name).strip() if customer_name else "Unknown Customer"
                    property_name_str = str(property_name).strip()
                    address_str = str(address).strip() if address else ""
                    
                    # 1. Get or Create Customer
                    customer_id = customer_cache.get(customer_name_str)
                    if not customer_id:
                        result = await session.execute(
                            select(CustomerModel).where(CustomerModel.short_name == customer_name_str)
                        )
                        customer = result.scalars().first()
                        if not customer:
                            customer = CustomerModel(
                                full_name=customer_name_str,
                                short_name=customer_name_str[:100],
                            )
                            session.add(customer)
                            await session.flush()
                            count_customers += 1
                        customer_id = customer.id
                        customer_cache[customer_name_str] = customer_id

                    # 2. Create Genba
                    # Check if Genba exists using memory cache
                    if property_name_str not in existing_genba_names:
                        genba = GenbaModel(
                            customer_id=customer_id,
                            property_name=property_name_str,
                            address=address_str,
                            status="ACTIVE",
                        )
                        session.add(genba)
                        existing_genba_names.add(property_name_str)
                        count_genbas += 1

                except Exception as e:
                    logger.warning(f"Error at row {row_idx}: {e}")
                    continue
            
            await session.commit()
            logger.info(f"Import completed. Created {count_customers} new customers and {count_genbas} new genbas.")
    except Exception as e:
        logger.error(f"Fatal error during import: {e}")
    finally:
        await engine.dispose()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python -m app.scripts.import_seed_data /path/to/excel.xlsx")
        sys.exit(1)
    asyncio.run(import_data(sys.argv[1]))
