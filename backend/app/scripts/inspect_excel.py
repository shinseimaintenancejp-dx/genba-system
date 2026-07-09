import sys
import openpyxl

def main(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        print(f"Sheet: {sheet_name}")
        ws = wb[sheet_name]
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            print(row)
        print("-" * 50)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        main(sys.argv[1])
    else:
        print("Provide file path")
